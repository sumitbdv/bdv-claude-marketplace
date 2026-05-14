#!/usr/bin/env python3
"""Read a single lab tab from the BDV taxonomy and emit normalised JSON.

Usage:
    python3 read_taxonomy.py --lab Skin
    python3 read_taxonomy.py --lab Wellness --status "Not Started"
"""
import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl missing. Run: pip3 install openpyxl\n")
    sys.exit(2)

LAB_TABS = {
    "skin": "SKIN LAB",
    "face": "FACE LAB",
    "hair": "HAIR LAB",
    "body": "BODY LAB",
    "wellness": "WELLNESS LAB",
}

HEADER_TOKENS = ("sub-section", "topic / modality", "topic/modality")


def find_data_path() -> Path:
    root = os.environ.get("CLAUDE_PLUGIN_ROOT")
    if root:
        p = Path(root) / "data" / "taxonomy.xlsx"
        if p.exists():
            return p
    here = Path(__file__).resolve().parent.parent / "data" / "taxonomy.xlsx"
    if here.exists():
        return here
    sys.stderr.write(f"taxonomy.xlsx not found (looked under CLAUDE_PLUGIN_ROOT and {here})\n")
    sys.exit(2)


def parse_lab(ws):
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        joined = " ".join("" if c is None else str(c).strip().lower() for c in row)
        if any(tok in joined for tok in HEADER_TOKENS):
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = ["" if c is None else str(c).strip() for c in rows[header_idx]]

    def col(name):
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i
        return -1

    idx = {
        "sub_section": col("Sub-Section"),
        "topic": col("Topic / Modality"),
        "mechanism": col("Clinical Mechanism"),
        "sub_topics": col("Sub-Topics"),
        "opening_hook": col("Opening Hook"),
        "dr_vali_notes": col("Dr Vali Notes"),
        "creator_1": col("Creator 1 Episode"),
        "creator_2": col("Creator 2 Episode"),
        "creator_3": col("Creator 3 Episode"),
        "transcript": col("Transcript Link"),
        "status": col("Status"),
        "podcast_script": col("Podcast Script"),
        "insta_script": col("Insta Script"),
        "tiktok_script": col("TikTok Script"),
        "approved": col("Approved"),
    }

    out = []
    for row in rows[header_idx + 1:]:
        if idx["sub_section"] == -1 or idx["topic"] == -1:
            continue
        sub_section = row[idx["sub_section"]] if idx["sub_section"] < len(row) else None
        topic = row[idx["topic"]] if idx["topic"] < len(row) else None
        if not topic or str(topic).strip() == "":
            continue
        if sub_section and str(sub_section).strip().upper() in {"LLAS", "PLAL"}:
            # section divider row
            continue

        def get(key):
            i = idx[key]
            if i < 0 or i >= len(row):
                return ""
            v = row[i]
            return "" if v is None else str(v).strip()

        entry = {
            "sub_section": get("sub_section"),
            "topic": get("topic"),
            "mechanism": get("mechanism"),
            "sub_topics": get("sub_topics"),
            "opening_hook": get("opening_hook"),
            "dr_vali_notes": get("dr_vali_notes"),
            "creators": [c for c in (get("creator_1"), get("creator_2"), get("creator_3")) if c],
            "transcript_link": get("transcript"),
            "status": get("status") or "Not Started",
            "approved": bool(get("approved")),
        }
        out.append(entry)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--lab", required=True, help="Skin | Face | Hair | Body | Wellness")
    ap.add_argument("--status", help="Filter by status (e.g. 'Not Started', 'Script Drafted')")
    args = ap.parse_args()

    lab_key = args.lab.strip().lower()
    if lab_key not in LAB_TABS:
        sys.stderr.write(f"Unknown lab '{args.lab}'. Choose from: {', '.join(LAB_TABS)}\n")
        sys.exit(2)

    wb = openpyxl.load_workbook(find_data_path(), data_only=True)
    sheet_name = LAB_TABS[lab_key]
    if sheet_name not in wb.sheetnames:
        sys.stderr.write(f"Sheet '{sheet_name}' not in workbook\n")
        sys.exit(2)

    rows = parse_lab(wb[sheet_name])
    if args.status:
        s = args.status.lower()
        rows = [r for r in rows if r["status"].lower() == s]
    json.dump({"lab": sheet_name, "count": len(rows), "rows": rows}, sys.stdout, indent=2, ensure_ascii=False)


if __name__ == "__main__":
    main()
